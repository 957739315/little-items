import openpyxl

class CorpusOutput(object):
    def write_excel(self,excel_name,lang_contents):
        #必须有两种语言以上
        if lang_contents is None:
            return False
        else:
            try:
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.title = excel_name


                result=self.deal_align(lang_contents)
                if result is None:
                    return False
                max_item=result[0]
                lang_contents=result[1]

                # 写入表头
                lang_name=list(lang_contents.keys())  #python3中dict.keys()返回是个set()
                for col in range(1, len(lang_name) + 1):
                    sheet.cell(row=1, column=col, value=lang_name[col - 1])

                values=list(lang_contents.values())
                for row in range(2,max_item+2):
                    for col in range(1,len(lang_name)+1):
                        sheet.cell(row=row, column=col, value=values[col-1][row-2])

                wb.save(excel_name+".xlsx")
                print(excel_name+"写入数据成功！\n")
                return True
            except Exception as e:
                print("write_excel exception!!"+str(e)+"     "+excel_name)
                return False




    def deal_align(self,lang_contents):
        '''处理对齐'''
        if lang_contents is None:
            return None
        max_item = 0

        flag=[]
        #删除没有语料的语种
        for key,value in lang_contents.items():
            if value is None:
                flag.append(key)
            else:
                if max_item < len(value):
                    max_item = len(value)

        '''
        #如果有存在空的语种，仍然写入
        #删除空语种，不能在遍历的时候删
        for f in flag:
            del lang_contents[f]
        '''

        ##如果有存在空的语种，则不写入写入
        if len(flag)>0:
            return None


        #存在的语料获取到的句子数不相同，则进行不写入
        k=0
        for key,value in lang_contents.items():
            while len(value)>max_item:
                k+=1
        if k>2:
            print("句子数对不齐")
            return None


        return max_item,lang_contents




